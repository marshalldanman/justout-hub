#!/usr/bin/env python3
"""
Gegenkraft Google Sheets Generator
===================================
Generates a fully formatted .xlsx file with all Gegenkraft data.
Open it directly in Google Sheets — no API credentials needed.

Usage:
  python3 generate_gegenkraft_sheet.py [--start-date 2026-03-30] [--days 30]

Output:
  gegenkraft/Gegenkraft_GK_Power_Dashboard.xlsx

Requirements:
  pip install openpyxl
"""

import argparse
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ═══════════════════════════════════════════════════════════════
#  STYLE CONSTANTS
# ═══════════════════════════════════════════════════════════════

HEADER_FILL = PatternFill(start_color="1F1F1F", end_color="1F1F1F", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
ALT_ROW_FILL = PatternFill(start_color="F2F2F5", end_color="F2F2F5", fill_type="solid")
WHITE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
BODY_FONT = Font(name="Calibri", size=10)
WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    bottom=Side(style="thin", color="DDDDDD"),
    right=Side(style="thin", color="EEEEEE"),
)

# Level colors for conditional formatting
LEVEL_COLORS = {
    "Recruit": "BDC3C7",
    "Resister": "85C1E9",
    "Disruptor": "F9E79F",
    "Saboteur": "F0B27A",
    "Breaker": "E74C3C",
    "Gegenkraft Master": "8E44AD",
}


# ═══════════════════════════════════════════════════════════════
#  DATA
# ═══════════════════════════════════════════════════════════════

DEFINITIONS = {
    "title": "Definitions",
    "headers": ["Term", "Definition", "Aliases", "Core Principle"],
    "col_widths": [20, 70, 35, 25],
    "rows": [
        [
            "Gegenkraft",
            "The action-power of intentional daily opposition of a habit \u2014 the intentional wedge that purposely sheds bad habits through deliberate denial/replacement, strategical countering or sabotage; its effectiveness tied to the law of momentum/inertia.",
            "Opposing Power, Gk-power, gkpower, gk",
            "Law of Momentum/Inertia",
        ],
        ["GK Move", "A specific daily action designed to oppose and disrupt a target habit", "Counter-move", "Deliberate opposition"],
        ["GK Power Level", "Self-rated daily score (1-10) of one's resistance strength and commitment", "Power Level", "Self-awareness"],
        ["Boss Battle", "A day with 3 or more urges \u2014 surviving clean earns double points", "High-urge day", "Resilience under pressure"],
        ["GK Streak", "Consecutive days of completed morning declarations and logged resistance", "Streak", "Momentum building"],
    ],
}

GK_MOVES = {
    "title": "GK Moves - Gambling",
    "headers": ["#", "Move Name", "Description", "Category", "Daily Action", "Difficulty", "Impact"],
    "col_widths": [5, 22, 55, 14, 35, 12, 10],
    "rows": [
        [1, "Delete & Block", "Uninstall every gambling app and use phone content restrictions to block re-downloading. Daily check: if it's back, it's gone again.", "Prevention", "Check app store restrictions are active", "Easy", "High"],
        [2, "Kill the Funding", "Remove saved payment methods, cards, and autofill credentials tied to gambling sites. Re-entering info = friction wall.", "Prevention", "Verify no saved payment methods exist", "Easy", "High"],
        [3, "Redirect the Trigger", "Identify the exact moment you reach for a gambling app (boredom/downtime/stress) and replace with a 5-minute action: walk, pushups, language lesson, journaling.", "Replacement", "Perform replacement action when triggered", "Medium", "High"],
        [4, "Screen Time Sabotage", "Set daily screen time limit of 0 minutes on gambling-category apps/websites. Someone else holds the Screen Time passcode.", "Prevention", "Confirm screen time limits are enforced", "Easy", "High"],
        [5, "Environment Poisoning", "Change phone wallpaper to bank statement showing losses or a written reminder of why you're stopping.", "Psychological", "Keep wallpaper as anti-gambling reminder", "Easy", "Medium"],
        [6, "Announce the Bet You Didn't Make", "Each day write down one specific bet you would have made and the amount. Track running total of money saved.", "Tracking", "Log the bet not made and amount saved", "Medium", "High"],
        [7, "Notification Detox", "Unsubscribe from every gambling promo email, block SMS marketing, turn off push notifications.", "Prevention", "Check for and block any new gambling notifications", "Easy", "Medium"],
        [8, "Accountability Ping", "Text one trusted person daily: 'Day X, no gambling.' Social commitment creates cost to relapsing.", "Social", "Send daily accountability text", "Medium", "High"],
        [9, "Cash the Urge", "When urge hits, immediately transfer the amount you would have gambled into a separate savings account.", "Financial", "Transfer would-be-bet money to savings", "Medium", "High"],
        [10, "Time Audit Log", "At end of each day log what you did with the time you would have spent gambling. Even 15 min documented.", "Tracking", "Log time reclaimed and activity done", "Easy", "Medium"],
    ],
}

BEE_INTEGRATION = {
    "title": "Bee AI x GK Integration",
    "headers": ["#", "Name", "Bee Feature", "GK Application", "Points", "How It Works"],
    "col_widths": [5, 25, 25, 20, 22, 60],
    "rows": [
        [1, "Voice Note Confessionals", "Voice Notes", "Urge Logging", "+5 per urge resisted", "Press Bee button when urge hits. Speak: 'GK Log: urge at [time], intensity [1-10], triggered by [cause].'"],
        [2, "Daily Insights Scoreboard", "Daily Insights", "Pattern Tracking", "+10 per week trend drops", "Bee surfaces emotional patterns. Track how often gambling-related language appears. Declining frequency = GK momentum proof."],
        [3, "Accountability Auto-Summary", "Conversation Segmentation", "Accountability Records", "+3 per conversation", "Wear Bee during check-ins with accountability partner. Bee auto-summarizes commitments made."],
        [4, "GK Daily Report Template", "Templates", "Daily Scorecard", "N/A (enables scoring)", "Custom 'Gegenkraft Daily Report' template: Urges resisted, Replacement actions, Money saved, GK power level."],
        [5, "Trigger Geo-Fencing", "Geo/Topic Fencing (future)", "Topic Boundaries", "-5 penalty", "Set topic boundaries around gambling language. If gambling topics appear in Bee themes, flag and activate counter-move."],
        [6, "Auto-Save the Bet Money", "Actions (Email/Calendar)", "Financial Redirection", "+1 per dollar redirected", "Say 'I just resisted a $50 bet' and Bee drafts email/reminder to transfer that amount to savings."],
        [7, "GK Boss Battles", "Developer API", "Scoring Engine", "Double points", "API extension pulls daily urge logs, calculates GK Power Score, flags 'boss battle' days (3+ urges), awards double points if survived clean."],
        [8, "Morning Declarations", "Voice Notes", "Daily Intention", "+2 (+streak multiplier)", "Each morning: 'Day X. No gambling today. GK power level: [1-10].' x1.5 after 7 days, x2 after 30."],
        [9, "Relationship Pattern Tracking", "Daily Insights (Relationships)", "Social Change Tracking", "+15 per positive shift", "Bee tracks relationship shifts. Monitor if gambling conversations with friends decrease. Quantified social change."],
        [10, "Weekly GK Retrospective", "Accumulated Summaries", "Weekly Review", "+20 per review", "Review Bee's weekly summaries. Count wins, flag weakest trigger points, set next week's GK targets."],
    ],
}

SCORING_RULES = {
    "title": "Scoring Rules",
    "headers": ["Action", "Points", "Condition", "Category"],
    "col_widths": [45, 10, 45, 18],
    "rows": [
        ["Log an urge resisted (via Bee voice note)", "+5", "Per urge logged and not acted on", "Urge Management"],
        ["Weekly gambling-language trend drops", "+10", "Per week the Bee insights trend line declines", "Pattern Tracking"],
        ["Accountability conversation captured", "+3", "Per conversation recorded via Bee", "Social"],
        ["Dollar redirected to savings", "+1", "Per dollar transferred instead of gambled", "Financial"],
        ["Morning GK declaration", "+2", "Per morning declaration via Bee voice note", "Intention Setting"],
        ["Morning declaration streak (7+ days)", "x1.5", "Multiplier applied to morning declaration points", "Streak Bonus"],
        ["Morning declaration streak (30+ days)", "x2.0", "Multiplier applied to morning declaration points", "Streak Bonus"],
        ["Positive relationship shift flagged by Bee", "+15", "When Bee surfaces declining gambling social patterns", "Social"],
        ["Weekly GK retrospective completed", "+20", "Per completed weekly review of Bee summaries", "Review"],
        ["Boss battle survived (3+ urges in one day)", "x2.0", "Double points for all actions that day", "Boss Battle"],
        ["Gambling topics appear in Bee themes", "-5", "Penalty per occurrence", "Penalty"],
    ],
}

LEVELS = {
    "title": "Level Progression",
    "headers": ["Level", "GK Points Required", "Title", "Description"],
    "col_widths": [8, 20, 22, 50],
    "rows": [
        [1, "0\u201350", "Recruit", "Just starting the fight. Building awareness and first defenses."],
        [2, "51\u2013150", "Resister", "Actively resisting urges. Habits are being challenged daily."],
        [3, "151\u2013300", "Disruptor", "Disrupting the pattern. Replacement behaviors are taking hold."],
        [4, "301\u2013500", "Saboteur", "Sabotaging the old habit loop. Momentum is shifting."],
        [5, "501\u2013750", "Breaker", "The habit's grip is breaking. New identity forming."],
        [6, "751+", "Gegenkraft Master", "Full opposing power achieved. The habit has lost its hold."],
    ],
}


# ═══════════════════════════════════════════════════════════════
#  SHEET BUILDER
# ═══════════════════════════════════════════════════════════════

def style_sheet(ws, sheet_data):
    """Apply headers, data, formatting, and column widths to a worksheet."""
    headers = sheet_data["headers"]
    rows = sheet_data["rows"]
    col_widths = sheet_data.get("col_widths", [20] * len(headers))

    # Column widths
    for i, width in enumerate(col_widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = width

    # Header row
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    # Data rows
    for row_idx, row_data in enumerate(rows, 2):
        fill = ALT_ROW_FILL if row_idx % 2 == 0 else WHITE_FILL
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = BODY_FONT
            cell.fill = fill
            cell.alignment = WRAP_ALIGN
            cell.border = THIN_BORDER

    # Freeze header row
    ws.freeze_panes = "A2"


def build_tracker_sheet(ws, start_date: str, days: int):
    """Build the daily tracker with formulas and conditional formatting."""
    headers = [
        "Date", "Day #", "Morning Declaration\n(Y/N)", "Urges\nResisted",
        "Urge Intensity\n(avg 1-10)", "Replacement\nActions Taken",
        "Money\nSaved ($)", "Accountability\nPing (Y/N)",
        "Bee Conversations\nCaptured", "Boss Battle\nDay (Y/N)",
        "GK Points\nEarned", "Cumulative\nPoints", "GK Level", "Notes",
    ]
    col_widths = [13, 7, 16, 10, 13, 15, 11, 15, 16, 13, 12, 13, 18, 30]

    # Column widths
    for i, width in enumerate(col_widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = width

    # Header row
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    # Row height for header
    ws.row_dimensions[1].height = 35

    start = datetime.strptime(start_date, "%Y-%m-%d")

    for i in range(days):
        row = i + 2
        day = start + timedelta(days=i)
        fill = ALT_ROW_FILL if row % 2 == 0 else WHITE_FILL

        # Date (col A)
        ws.cell(row=row, column=1, value=day.strftime("%Y-%m-%d")).font = BODY_FONT
        # Day # (col B)
        ws.cell(row=row, column=2, value=i + 1).font = BODY_FONT

        # Empty input columns C-J, K
        for col in range(3, 12):
            ws.cell(row=row, column=col).font = BODY_FONT

        # Cumulative Points formula (col L = 12)
        if i == 0:
            ws.cell(row=row, column=12, value=f"=K{row}").font = BODY_FONT
        else:
            ws.cell(row=row, column=12, value=f"=L{row - 1}+K{row}").font = BODY_FONT

        # GK Level formula (col M = 13)
        level_formula = (
            f'=IF(L{row}>=751,"Gegenkraft Master",'
            f'IF(L{row}>=501,"Breaker",'
            f'IF(L{row}>=301,"Saboteur",'
            f'IF(L{row}>=151,"Disruptor",'
            f'IF(L{row}>=51,"Resister","Recruit")))))'
        )
        ws.cell(row=row, column=13, value=level_formula).font = Font(name="Calibri", size=10, bold=True)

        # Notes (col N = 14)
        ws.cell(row=row, column=14).font = BODY_FONT

        # Apply row fill and borders
        for col in range(1, 15):
            cell = ws.cell(row=row, column=col)
            cell.fill = fill
            cell.border = THIN_BORDER
            if col in (2, 3, 4, 5, 8, 10):
                cell.alignment = CENTER_ALIGN
            else:
                cell.alignment = WRAP_ALIGN

    # Freeze header
    ws.freeze_panes = "A2"


# ═══════════════════════════════════════════════════════════════
#  MAIN
# ═══════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(description="Generate Gegenkraft Excel/Google Sheets file")
    parser.add_argument("--start-date", default=datetime.now().strftime("%Y-%m-%d"), help="Start date (YYYY-MM-DD)")
    parser.add_argument("--days", type=int, default=30, help="Number of days to track")
    args = parser.parse_args()

    wb = Workbook()

    # ── Sheet 1: Definitions ──
    ws1 = wb.active
    ws1.title = "Definitions"
    style_sheet(ws1, DEFINITIONS)

    # ── Sheet 2: GK Moves ──
    ws2 = wb.create_sheet("GK Moves - Gambling")
    style_sheet(ws2, GK_MOVES)

    # ── Sheet 3: Bee Integration ──
    ws3 = wb.create_sheet("Bee AI x GK Integration")
    style_sheet(ws3, BEE_INTEGRATION)

    # ── Sheet 4: Scoring Rules ──
    ws4 = wb.create_sheet("Scoring Rules")
    style_sheet(ws4, SCORING_RULES)

    # ── Sheet 5: Levels ──
    ws5 = wb.create_sheet("Level Progression")
    style_sheet(ws5, LEVELS)

    # Color-code level titles
    for row_idx, row_data in enumerate(LEVELS["rows"], 2):
        title = row_data[2]
        if title in LEVEL_COLORS:
            ws5.cell(row=row_idx, column=3).fill = PatternFill(
                start_color=LEVEL_COLORS[title],
                end_color=LEVEL_COLORS[title],
                fill_type="solid",
            )
            if title in ("Breaker", "Gegenkraft Master"):
                ws5.cell(row=row_idx, column=3).font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")

    # ── Sheet 6: Daily Tracker ──
    ws6 = wb.create_sheet("Daily Tracker")
    build_tracker_sheet(ws6, args.start_date, args.days)

    # Save
    out_path = Path(__file__).parent / "Gegenkraft_GK_Power_Dashboard.xlsx"
    wb.save(str(out_path))

    print(f"""
========================================================
  GEGENKRAFT DASHBOARD GENERATED
========================================================

  File: {out_path}

  Tabs:
    1. Definitions           - GK terminology
    2. GK Moves - Gambling   - 10 counter-moves
    3. Bee AI x GK           - 10 Bee integrations
    4. Scoring Rules         - Point system & multipliers
    5. Level Progression     - Recruit -> Gegenkraft Master
    6. Daily Tracker         - {args.days}-day tracker with formulas

  Start date: {args.start_date}

  TO OPEN IN GOOGLE SHEETS:
    1. Go to sheets.google.com
    2. File -> Open -> Upload
    3. Select: Gegenkraft_GK_Power_Dashboard.xlsx
    4. Done. All 6 tabs, formatting, and formulas are ready.
========================================================
""")


if __name__ == "__main__":
    main()
